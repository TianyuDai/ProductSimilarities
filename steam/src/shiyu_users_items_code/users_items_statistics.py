import collections

import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import sklearn.manifold as manifold_learning
from sklearn.cluster import SpectralClustering
import pandas as pd

import steam.src.shiyu_users_items_code.config as config
import steam.src.shiyu_users_items_code.plotting as plotting

gzip_load = config.gzip_load
gzip_save = config.gzip_save


class UsersItemsStat(object):
    def __init__(self):
        self.execute_loop = False
        self.input_file_path = None
        self.items_counter = collections.Counter()
        self.game_id_buyers_index_dict = {}
        self.items_playtime_2weeks_counter = collections.Counter()
        self.items_playtime_forever_counter = collections.Counter()
        self.game_name_dict = {}
        self.game_num_to_stat = None
        self.complete_game_id_list = []
        self.id_item_dict = {}
        self.ignored_games_dict = None
        self.example_game_id_playtime_dict = {}
        self.game_player_playtime_dict = {}
        self.player_game_playtime_dict = {}
        self.person_playtime_pair = []
        self.player_weight_dict = None
        self.basic_playtime = 0
        self.game_similarity_matrix = None
        self.embedding_dim = -1
        self.cluster_num = -1
        self.max_item_num = -1
        self.transformed_coordinates = None
        self.game_id_list_training_set = None

    def hook_prepare(self):
        # self.input_file_path = config.test_users_items_file
        self.input_file_path = config.users_items_file
        complete_game_id_list = True
        id_item_dict = False
        game_id_buyer_index_dict = False
        game_id_list_training_set = False
        game_player_playtime_dict = False
        player_game_playtime_dict = False
        player_weight_dict = False
        game_similarity_matrix = True
        transformed_coordinates = False

        self.execute_loop = False
        self.game_num_to_stat = 500
        self.basic_playtime = 10
        self.example_game_id_playtime_dict = {'620': [], '236390': [], '24240': [], '12210': []}

        self.embedding_dim = 3
        self.max_item_num = 8500
        self.cluster_num = 8

        # self.ignored_games_dict = self.ignored_games_loader()
        self.pickle_obj_loader(
            complete_game_id_list, id_item_dict, game_id_buyer_index_dict, game_id_list_training_set,
            game_player_playtime_dict, player_game_playtime_dict, player_weight_dict, game_similarity_matrix,
            transformed_coordinates)

    def hook_each_line(self, input_dict):
        self.stat_purchase_and_playtime_each_line(input_dict)
        # self.game_player_playtime_dict_generator(input_dict)
        # self.player_game_playtime_dict_generator(input_dict)
        # self.player_playtime_distribution(input_dict)
        pass

    def hook_final(self):
        # self.complete_item_stat()
        # self.average_playtime_stat()
        # self.item_information_distribution_plot()
        # self.popular_game_information_distribution_plot()
        # self.parameter_saver()
        # self.total_playtime_each_person_plot()
        # self.player_weight_dict_saver()
        # self.game_index_stat()
        # self.purchase_similarity_computation()
        # self.similarity_plot()
        # self.similarity_embedding()
        # self.embedding_plot()
        # self.le_similarity_for_training_output()
        self.spectral_clustering()

    def pickle_obj_loader(
            self, complete_game_id_list=False, id_item_dict=False, game_id_buyer_index_dict=False,
            game_id_list_training_set=False, game_player_playtime_dict=False, player_game_playtime_dict=False,
            player_weight_dict=False, game_similarity_matrix=False, transformed_coordinates=False):

        if complete_game_id_list:
            self.complete_game_id_list = gzip_load(config.complete_game_id_list_file)
        if id_item_dict:
            self.id_item_dict = gzip_load(config.game_id_name_dict_file)
        if game_id_buyer_index_dict:
            self.game_id_buyers_index_dict = gzip_load(config.game_id_purchase_number_order_dict_file)
        if game_id_list_training_set:
            self.game_id_list_training_set = gzip_load(config.final_nn_data_game_id_list)
        if game_player_playtime_dict:
            self.game_player_playtime_dict = gzip_load(config.game_player_playtime_dict_file)
        if player_game_playtime_dict:
            self.player_game_playtime_dict = gzip_load(config.player_game_playtime_dict_file)
        if player_weight_dict:
            self.player_weight_dict = gzip_load(config.player_weight_dict_file)
        if game_similarity_matrix:
            self.game_similarity_matrix = config.np_load(
                config.game_weighted_similarity_matrix, 'game_similarity_matrix')
        if transformed_coordinates:
            self.transformed_coordinates = config.np_load(
                "{}_{}d.npz".format(config.embedded_coordinates_laplacian_eigenmaps, self.embedding_dim),
                config.output_data_matrix_name)

    def stat_purchase_and_playtime_each_line(self, input_dict):
        for item in input_dict['items']:
            item_name = item['item_name']
            item_id = item['item_id']
            self.id_item_dict[item_id] = item_name

            self.items_counter[item_id] += 1
            self.items_playtime_2weeks_counter[item_id] += item['playtime_2weeks']
            self.items_playtime_forever_counter[item_id] += item['playtime_forever']

            if item_id in self.example_game_id_playtime_dict:
                self.example_game_id_playtime_dict[item_id].append(item['playtime_forever'])

    def game_player_playtime_dict_generator(self, input_dict):
        steam_id = input_dict['steam_id']
        game_player_playtime_dict = self.game_player_playtime_dict
        for item in input_dict['items']:
            item_name = item['item_name']
            item_id = item['item_id']
            self.id_item_dict[item_id] = item_name
            if item_id not in game_player_playtime_dict:
                game_player_playtime_dict[item_id] = {}
            current_item_dict = game_player_playtime_dict[item_id]
            if steam_id not in current_item_dict:
                current_item_dict[steam_id] = 0
            current_item_dict[steam_id] += item['playtime_forever']

    def player_game_playtime_dict_generator(self, input_dict):
        steam_id = input_dict['steam_id']
        current_game_id_playtime_dict = {}
        for item in input_dict['items']:
            item_name = item['item_name']
            item_id = item['item_id']
            self.id_item_dict[item_id] = item_name
            current_game_id_playtime_dict[item_id] = item['playtime_forever']
        self.player_game_playtime_dict[steam_id] = current_game_id_playtime_dict

    def player_playtime_distribution(self, input_dict):
        total_playtime = 0
        steam_id = input_dict['steam_id']
        for item in input_dict['items']:
            total_playtime += item['playtime_forever']
        # self.playtime_each_person.append(total_playtime)
        self.person_playtime_pair.append((steam_id, total_playtime))

    @staticmethod
    def ignored_games_loader():
        ignored_games_dict = {}
        wb = openpyxl.load_workbook(config.ignored_games_list_file)
        for ignored_games_worksheet in wb.worksheets:
            for game_id_cell, game_name_cell in ignored_games_worksheet.rows:
                game_id = game_id_cell.value
                game_name = game_name_cell.value
                if game_id == config.game_id_label:
                    continue
                ignored_games_dict[game_id] = game_name
        return ignored_games_dict

    @staticmethod
    def insert_or_refresh_sheet(current_sheet_name):
        try:
            wb = openpyxl.load_workbook(config.game_stat_file)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
        try:
            del wb[current_sheet_name]
        except KeyError:
            pass
        worksheet = wb.create_sheet(current_sheet_name)
        return worksheet, wb

    def complete_item_stat(self):
        total_game_num = self.game_num_to_stat
        current_sheet_name = 'basic_stat'
        worksheet, wb = self.insert_or_refresh_sheet(current_sheet_name)
        worksheet.append((
            config.game_name_label, config.game_buyers_label, "",
            config.game_name_label, config.game_two_weeks_playtime_label, "",
            config.game_name_label, config.game_total_playtime_label))
        for ((item_id, purchase_num),
             (item_id_playtime_2weeks, playtime_2weeks),
             (item_id_playtime_forever, playtime_forever)) in zip(
                self.items_counter.most_common(total_game_num),
                self.items_playtime_2weeks_counter.most_common(total_game_num),
                self.items_playtime_forever_counter.most_common(total_game_num)):
            worksheet.append((
                self.id_item_dict[item_id], purchase_num, "",
                self.id_item_dict[item_id_playtime_2weeks], playtime_2weeks, "",
                self.id_item_dict[item_id_playtime_forever], playtime_forever))
        wb.save(config.game_stat_file)

    def game_index_stat(self):
        self.game_id_buyers_index_dict = {
            game_id: index for index, (game_id, _) in enumerate(self.items_counter.most_common())}
        gzip_save(self.game_id_buyers_index_dict, config.game_id_purchase_number_order_dict_file)

    def average_playtime_stat(self):
        current_sheet_name = 'average_playtime'
        worksheet, wb = self.insert_or_refresh_sheet(current_sheet_name)
        worksheet.append((
            config.game_id_label, config.game_name_label, config.game_buyers_label,
            config.game_average_two_weeks_playtime_label,
            config.game_average_total_playtime_label))
        for item_id, purchase_num in self.items_counter.most_common(self.game_num_to_stat):
            two_weeks_playtime = self.items_playtime_2weeks_counter[item_id]
            forever_playtime = self.items_playtime_forever_counter[item_id]
            two_weeks_playtime_each = two_weeks_playtime / purchase_num
            forever_playtime_each = forever_playtime / purchase_num
            worksheet.append((
                item_id, self.id_item_dict[item_id], purchase_num, two_weeks_playtime_each, forever_playtime_each))
        wb.save(config.game_stat_file)

    def item_information_distribution_plot(self):
        data_array = np.array(
            [purchase_num for item_id, purchase_num in self.items_counter.most_common()])
        plotting.bar_histogram_plot(data_array, 'Purchase number', x_lim=[0, 5000], y_lim=[0, 20000])
        data_array = np.array(
            [playtime_forever for item_id, playtime_forever in self.items_playtime_forever_counter.most_common()])
        plotting.bar_histogram_plot(data_array, 'Playtime forever', x_lim=[0, 5000], y_lim=[0, 1e7])
        plt.show()

    def popular_game_information_distribution_plot(self):
        for item_id, playtime_list in self.example_game_id_playtime_dict.items():
            item_name = self.id_item_dict[item_id]
            data_array = -np.sort(-np.array(playtime_list))
            plotting.bar_histogram_plot(data_array, 'Playtime forever: {}'.format(item_name), x_lim=[0, 1e4], y_lim=[0, 1e4])
        plt.show()

    def total_playtime_each_person_plot(self):
        data_array = -np.sort(-np.array(list(zip(*self.person_playtime_pair))[1]))
        plotting.bar_histogram_plot(data_array, 'Total playtime for each person', x_lim=[0, 1e5], y_lim=[0, 3e5])
        print(len(data_array))
        plt.show()

    def similarity_plot(self):
        data_array = self.game_similarity_matrix
        # data_array -= np.diag(np.diag(data_array))
        data_vector = np.reshape(data_array, [-1])
        data_vector_sample = np.random.choice(data_vector, int(1e5))
        plotting.violin_plot(data_vector_sample, [0, 0.001], label='Similarity')
        # plotting.heatmap_plot(data_array, 'Game similarity', max_value=0.0008)
        plt.show()

    def embedding_plot(self):
        transformed_coordinates = self.transformed_coordinates
        dim = self.embedding_dim
        color_rgb = np.array([31, 119, 180]) / 255
        alpha_value = 0.2
        x_lim = [-0.035, 0.035]
        y_lim = [-0.06, 0.105]
        color_vector = np.concatenate([color_rgb, [alpha_value]])
        if dim == 2:
            plotting.scatter2d_plot(
                transformed_coordinates, marker_size=3, color=color_vector, x_lim=x_lim, y_lim=y_lim)
        elif dim == 3:
            plotting.scatter3d_plot(
                transformed_coordinates, marker_size=3, color=color_vector, x_lim=x_lim, y_lim=y_lim)
        if x_lim[0] is not None:
            outlier_loc = transformed_coordinates[:, 0] < x_lim[0]
            print(np.count_nonzero(outlier_loc) / transformed_coordinates.shape[0])
        if x_lim[1] is not None:
            outlier_loc = transformed_coordinates[:, 0] > x_lim[1]
            print(np.count_nonzero(outlier_loc) / transformed_coordinates.shape[0])
        if y_lim[0] is not None:
            outlier_loc = transformed_coordinates[:, 1] < y_lim[0]
            print(np.count_nonzero(outlier_loc) / transformed_coordinates.shape[0])
        if y_lim[1] is not None:
            outlier_loc = transformed_coordinates[:, 1] > y_lim[1]
            print(np.count_nonzero(outlier_loc) / transformed_coordinates.shape[0])
        plt.show()

    def parameter_saver(self):
        gzip_save(self.id_item_dict, config.game_id_name_dict_file)
        gzip_save(self.game_player_playtime_dict, config.game_player_playtime_dict_file)
        gzip_save(self.player_game_playtime_dict, config.player_game_playtime_dict_file)

    def player_weight_dict_saver(self):
        sorted_person_playtime_pair = sorted(self.person_playtime_pair, key=lambda x: x[1])
        playtime_set = set()
        for _, playtime in sorted_person_playtime_pair:
            playtime_set.add(playtime)
        playtime_value_num = len(playtime_set)
        print(
            "Person number: {}\nPlaytime value number: {}".format(
                len(sorted_person_playtime_pair), playtime_value_num))

        player_weight_dict = {}
        current_playtime = 0
        current_weight = 1 / playtime_value_num
        for steam_id, playtime in sorted_person_playtime_pair:
            if playtime > current_playtime:
                current_playtime = playtime
                current_weight += 1 / playtime_value_num
            player_weight_dict[steam_id] = current_weight

        gzip_save(player_weight_dict, config.player_weight_dict_file)

    def purchase_similarity_computation(self):
        # game_id_name_dict = self.id_item_dict
        # game_id_index_dict = {game_id: index for index, game_id in enumerate(game_id_name_dict.keys())}
        game_id_buyers_index_dict = self.game_id_buyers_index_dict
        player_game_playtime_dict = self.player_game_playtime_dict
        game_num = len(game_id_buyers_index_dict)
        player_num = len(player_game_playtime_dict)
        raw_similarity_matrix = np.zeros([game_num, game_num])
        weight_matrix = np.ones([game_num, game_num]) * 1e-8
        # weight_matrix = np.zeros([game_num, game_num])
        player_weight_dict = self.player_weight_dict
        count = 0
        for player_id, game_playtime_dict in player_game_playtime_dict.items():
            player_weight = player_weight_dict[player_id]
            game_num = len(game_playtime_dict)
            game_id_list = list(game_playtime_dict.keys())
            game_index_list = [game_id_buyers_index_dict[game_id] for game_id in game_id_list]
            playtime_array = np.array(list(game_playtime_dict.values())) + self.basic_playtime
            total_playtime = playtime_array.sum()
            playtime_ratio_array = playtime_array / total_playtime
            for index_i in range(game_num):
                for index_j in range(index_i + 1, game_num):
                    playtime_ratio_i = playtime_ratio_array[index_i]
                    playtime_ratio_j = playtime_ratio_array[index_j]
                    game_index_i = game_index_list[index_i]
                    game_index_j = game_index_list[index_j]
                    weighted_similarity = (
                            playtime_ratio_i * playtime_ratio_j / (playtime_ratio_i + playtime_ratio_j)
                            * player_weight)
                    raw_similarity_matrix[game_index_i, game_index_j] += weighted_similarity
                    raw_similarity_matrix[game_index_j, game_index_i] += weighted_similarity
                    weight_matrix[game_index_i, game_index_j] += player_weight
                    weight_matrix[game_index_j, game_index_i] += player_weight
            count += 1
            if count % 1000 == 0:
                print("{:.3f} finished...".format(count / player_num))
        game_similarity_matrix = raw_similarity_matrix / weight_matrix
        np.savez_compressed(config.game_weighted_similarity_matrix, game_similarity_matrix=game_similarity_matrix)
        self.game_similarity_matrix = game_similarity_matrix

    def similarity_embedding(self):
        dim = self.embedding_dim
        spectral_embedding = manifold_learning.SpectralEmbedding(
            n_components=dim, affinity='precomputed', n_jobs=7)

        max_item_num = self.max_item_num
        similarity_matrix = self.game_similarity_matrix[0:max_item_num, 0:max_item_num]

        transformed_coordinates = spectral_embedding.fit_transform(similarity_matrix)
        np.savez_compressed(
            "{}_{}d".format(config.embedded_coordinates_laplacian_eigenmaps, dim),
            transformed_coordinates=transformed_coordinates)
        self.transformed_coordinates = transformed_coordinates

    def le_similarity_for_training_output(self):
        game_id_buyers_index_dict = self.game_id_buyers_index_dict
        transformed_coordinates = self.transformed_coordinates
        game_id_list = []
        current_game_index_list = []
        for game_id in self.game_id_list_training_set:
            current_game_index_list.append(game_id_buyers_index_dict[game_id])
            game_id_list.append(game_id)
        le_coordinates_for_training = transformed_coordinates[current_game_index_list, :]
        le_coordinates_dataframe = pd.DataFrame(le_coordinates_for_training, index=game_id_list)
        le_coordinates_dataframe.to_excel(config.final_game_le_similarity_output_df)

    def spectral_clustering(self):
        game_similarity_matrix = self.game_similarity_matrix
        complete_game_id_list = self.complete_game_id_list
        data_size = game_similarity_matrix.shape[0]
        spectral_clustering = SpectralClustering(
            n_clusters=self.cluster_num, affinity='precomputed', n_jobs=7)
        clustering_labels_array = spectral_clustering.fit_predict(game_similarity_matrix)
        counter = collections.Counter(clustering_labels_array)
        print(counter)
        np.savez_compressed(
            "{}".format(config.complete_cluster_labels_array),
            clustering_labels_array=clustering_labels_array)
        order_pair_list = [
            (index, index + label * 10 * data_size)
            for index, label in enumerate(clustering_labels_array)]
        new_order_array = list(zip(*sorted(order_pair_list, key=lambda x: x[1])))[0]
        clustered_similarity_matrix = shuffle_symmetric_matrix(game_similarity_matrix, new_order_array)
        plotting.heatmap_plot(clustered_similarity_matrix, max_value=0.0008)
        plt.show()

    # def cluster_feature_stat(self):



def example_data_generator():
    def generator_for_one_file(raw_file_name, example_file_name, count):
        with open(raw_file_name, encoding='utf-8') as f_in, open(example_file_name, 'w', encoding='utf-8') as f_out:
            for i in range(count):
                f_out.write(f_in.readline())

    line_num = 100
    generator_for_one_file(config.users_items_file, config.test_users_items_file, line_num)
    generator_for_one_file(config.steam_games_file, config.test_steam_games_file, line_num)


def json_file_loader(fp):
    for line in fp:
        yield eval(line)


def iterative_items_statistics(operation_obj):
    operation_obj.hook_prepare()
    if operation_obj.execute_loop:
        with open(operation_obj.input_file_path, encoding='utf-8') as f_in:
            for input_dict in json_file_loader(f_in):
                operation_obj.hook_each_line(input_dict)
    operation_obj.hook_final()


def symmetrize_matrix(matrix, upper_or_lower='upper'):
    if matrix.shape[0] != matrix.shape[1]:
        raise ValueError("Not symmetrical! Shape: {}".format(matrix.shape))
    new_matrix = np.copy(matrix)
    dim = new_matrix.shape[0]
    if upper_or_lower == 'upper':
        for i in range(dim):
            for j in range(i + 1, dim):
                new_matrix[j, i] = new_matrix[i, j]
    elif upper_or_lower == 'lower':
        for i in range(dim):
            for j in range(i + 1, dim):
                new_matrix[i, j] = new_matrix[j, i]
    else:
        raise ValueError('Parameter error! Only for "upper" or "lower"')
    return new_matrix


def shuffle_symmetric_matrix(input_matrix, target_order_list):
    if len(target_order_list) != input_matrix.shape[0] != input_matrix.shape[1]:
        raise ValueError('Length of target_order_list is not equal to dimension of input_matrix!')
    row_shuffled_matrix = input_matrix[target_order_list, :]
    col_shuffled_matrix = row_shuffled_matrix[:, target_order_list]
    return col_shuffled_matrix


def main():
    # example_data_generator()
    users_items_processor = UsersItemsStat()
    iterative_items_statistics(users_items_processor)


if __name__ == '__main__':
    main()
